"""CFTC 数据处理与 Excel 生成"""

import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from .config import (
    BACKTEST_SHEET, CATEGORIES, CFTC_XLS_FOLDER, DATE_COL,
    MARKET_COL, METAL_FILTERS, OI_COL, OUTPUT_EXCEL,
    OUTPUT_FOLDER, TIME_PERIODS,
)


# ====================== 处理单个文件 ======================
def process_file(file):
    try:
        df = pd.read_excel(file, engine='xlrd')
        if DATE_COL not in df.columns or MARKET_COL not in df.columns:
            return None
        df['Report_Date'] = pd.to_datetime(df[DATE_COL], errors='coerce')
        df = df.dropna(subset=['Report_Date'])

        mappings = {
            "Producers": ("Prod_Merc_Positions_Long_ALL", "Prod_Merc_Positions_Short_ALL"),
            "Swap_Dealers": ("Swap_Positions_Long_All", "Swap__Positions_Short_All"),
            "Money_Managers": ("M_Money_Positions_Long_ALL", "M_Money_Positions_Short_ALL"),
            "Other_Reportables": ("Other_Rept_Positions_Long_ALL", "Other_Rept_Positions_Short_ALL"),
            "Non_Reportables": ("NonRept_Positions_Long_All", "NonRept_Positions_Short_All")
        }

        for cat, (l, s) in mappings.items():
            if l in df.columns and s in df.columns:
                df[f"{cat}_Long"] = df[l]
                df[f"{cat}_Short"] = df[s]
                df[f"{cat}_Net"] = df[l] - df[s]
                df[f"{cat}_Pct_OI"] = df[f"{cat}_Net"] / df[OI_COL] * 100

        return df
    except Exception as e:
        print(f"处理文件失败: {e}")
        return None


# ====================== 主 sheet 处理 ======================
def process_and_write(writer, price_dfs):
    all_data = {m: pd.DataFrame() for m in METAL_FILTERS}

    for file in sorted(CFTC_XLS_FOLDER.glob("*.xls")):
        df = process_file(file)
        if df is not None:
            for metal, keywords in METAL_FILTERS.items():
                mask = df[MARKET_COL].str.upper().isin([k.upper() for k in keywords])
                if mask.any():
                    all_data[metal] = pd.concat([all_data[metal], df[mask]]).copy()

    for metal, df in all_data.items():
        if df.empty:
            print(f"{metal} 数据为空，跳过处理")
            continue

        df = df.sort_values('Report_Date').drop_duplicates('Report_Date', keep='last').copy()
        df = df[df['Report_Date'] >= '2007-01-01']
        df['Report_Date'] = pd.to_datetime(df['Report_Date'])
        df = df.set_index('Report_Date').copy()

        if metal in price_dfs and not price_dfs[metal].empty:
            price_series = price_dfs[metal]['close'].sort_index()
            df['Tuesday_Date'] = df.index - pd.Timedelta(days=3)
            df['close'] = np.nan
            for idx, row in df.iterrows():
                tue_date = row['Tuesday_Date']
                if tue_date in price_series.index:
                    df.at[idx, 'close'] = price_series.loc[tue_date]
                else:
                    prev_dates = price_series.index[price_series.index < tue_date]
                    if not prev_dates.empty:
                        nearest_date = prev_dates.max()
                        df.at[idx, 'close'] = price_series.loc[nearest_date]
            df['close'] = df['close'].ffill().bfill()

            print_check = df.reset_index()
            print(f"\n{metal} 最近10周 close 匹配检查：")
            print(print_check[['Report_Date', 'Tuesday_Date', 'close']].tail(10))

        df = df.reset_index()

        # 全历史分位
        for cat in CATEGORIES:
            pct_col = f"{cat}_Pct_OI"
            if pct_col not in df.columns:
                continue
            lower = df[pct_col].quantile(0.01)
            upper = df[pct_col].quantile(0.99)
            clipped = df[pct_col].clip(lower, upper)
            df[f"{cat}_Pct_OI_Percentile"] = clipped.rank(pct=True) * 100

        # 滚动窗口分位（不含Historical）
        latest = df['Report_Date'].max()
        for p_name, delta in TIME_PERIODS.items():
            p_df = df[df['Report_Date'] >= latest - delta].copy()
            for cat in CATEGORIES:
                pct_col = f"{cat}_Pct_OI"
                if pct_col in p_df.columns:
                    lower = p_df[pct_col].quantile(0.01)
                    upper = p_df[pct_col].quantile(0.99)
                    clipped = p_df[pct_col].clip(lower, upper)
                    ranks = clipped.rank(pct=True) * 100
                    df[f"{cat}_{p_name}_Pct_OI_Percentile"] = np.nan
                    df.loc[p_df.index, f"{cat}_{p_name}_Pct_OI_Percentile"] = ranks

        # 输出列
        cols = ['Report_Date', OI_COL]
        for cat in CATEGORIES:
            cols += [f"{cat}_Long", f"{cat}_Short", f"{cat}_Net", f"{cat}_Pct_OI", f"{cat}_Pct_OI_Percentile"]
            for p in TIME_PERIODS:
                cols.append(f"{cat}_{p}_Pct_OI_Percentile")
        if 'close' in df.columns:
            cols.append('close')

        df_output = df[cols].sort_values('Report_Date', ascending=False)
        df_output.to_excel(writer, sheet_name=metal, index=False, float_format="%.2f")

        # 颜色填充
        sheet = writer.sheets[metal]
        red = PatternFill("solid", start_color="FF0000")
        orange = PatternFill("solid", start_color="FFA500")
        blue = PatternFill("solid", start_color="0000FF")
        green = PatternFill("solid", start_color="00FF00")
        gray = PatternFill("solid", start_color="D3D3D3")

        for i, col in enumerate(df_output.columns):
            if 'Pct_OI_Percentile' not in col:
                continue
            letter = get_column_letter(i + 1)
            for r in range(2, len(df_output) + 2):
                cell = sheet[f"{letter}{r}"]
                try:
                    v = float(cell.value)
                    if v >= 80: cell.fill = red
                    elif 50 <= v < 80: cell.fill = orange
                    elif 20 <= v < 50: cell.fill = blue
                    elif v <= 20: cell.fill = green
                    else: cell.fill = gray
                except:
                    pass

        print(f"{metal} 主 sheet 已生成")

    return all_data


# ====================== 2014-2020 sheet ======================
def add_2014_2020_sheets(writer, all_data, price_dfs):
    for metal, raw_df in all_data.items():
        if raw_df.empty:
            print(f"{metal} 2014-2020 数据为空，跳过")
            continue

        df = raw_df.sort_values('Report_Date').drop_duplicates('Report_Date', keep='last').copy()
        df = df[(df['Report_Date'] >= '2014-01-01') & (df['Report_Date'] <= '2020-12-31')]
        if len(df) < 30:
            print(f"{metal} 2014-2020 数据不足，跳过")
            continue

        df = df.set_index('Report_Date')

        if metal in price_dfs and not price_dfs[metal].empty:
            price_series = price_dfs[metal]['close'].sort_index()
            df['Tuesday_Date'] = df.index - pd.Timedelta(days=3)
            df['close'] = np.nan
            for idx, row in df.iterrows():
                tue_date = row['Tuesday_Date']
                if tue_date in price_series.index:
                    df.at[idx, 'close'] = price_series.loc[tue_date]
                else:
                    prev_dates = price_series.index[price_series.index < tue_date]
                    if not prev_dates.empty:
                        nearest_date = prev_dates.max()
                        df.at[idx, 'close'] = price_series.loc[nearest_date]
            df['close'] = df['close'].ffill().bfill()

        df = df.reset_index()

        for cat in CATEGORIES:
            pct_col = f"{cat}_Pct_OI"
            if pct_col not in df.columns:
                continue
            lower = df[pct_col].quantile(0.01)
            upper = df[pct_col].quantile(0.99)
            clipped = df[pct_col].clip(lower, upper)
            df[f"{cat}_Pct_OI_Percentile"] = clipped.rank(pct=True) * 100

        if 'close' in df.columns and df['close'].notna().sum() > 1:
            df['Weekly_Return'] = df['close'].pct_change(1) * 100
            df['Next_Weekly_Return'] = df['close'].pct_change(1).shift(-1) * 100
        else:
            df['Weekly_Return'] = np.nan
            df['Next_Weekly_Return'] = np.nan

        for cat in CATEGORIES:
            pct_col = f"{cat}_Pct_OI_Percentile"
            if pct_col not in df.columns:
                continue
            df[f'{cat}_Signal'] = '中性'
            df.loc[df[pct_col] >= 80, f'{cat}_Signal'] = '极度拥挤多头'
            df.loc[(df[pct_col] >= 50) & (df[pct_col] < 80), f'{cat}_Signal'] = '强势多头 / 拥挤多头'
            df.loc[(df[pct_col] >= 20) & (df[pct_col] < 50), f'{cat}_Signal'] = '弱势空头 / 拥挤空头'
            df.loc[df[pct_col] < 20, f'{cat}_Signal'] = '极度拥挤空头'

            df[f'{cat}_Week_Match'] = 0
            df.loc[(df[pct_col] >= 50) & (df['Weekly_Return'] > 0), f'{cat}_Week_Match'] = 1
            df.loc[(df[pct_col] < 50) & (df['Weekly_Return'] < 0), f'{cat}_Week_Match'] = 1

            df[f'{cat}_Next_Match'] = 0
            df.loc[(df[pct_col] >= 50) & (df['Next_Weekly_Return'] > 0), f'{cat}_Next_Match'] = 1
            df.loc[(df[pct_col] < 50) & (df['Next_Weekly_Return'] < 0), f'{cat}_Next_Match'] = 1

        cols = ['Report_Date', OI_COL]
        if 'close' in df.columns:
            cols.append('close')
        for cat in CATEGORIES:
            cols += [f"{cat}_Long", f"{cat}_Short", f"{cat}_Net", f"{cat}_Pct_OI",
                     f"{cat}_Pct_OI_Percentile", f"{cat}_Signal", 
                     f"{cat}_Week_Match", f"{cat}_Next_Match"]
        cols += ['Weekly_Return', 'Next_Weekly_Return']

        df_output = df[cols].sort_values('Report_Date', ascending=False)
        sheet_name = f"{metal}_2014_2020"
        df_output.to_excel(writer, sheet_name=sheet_name, index=False, float_format="%.2f")

        # 颜色标记
        sheet = writer.sheets[sheet_name]
        red = PatternFill("solid", start_color="FF0000")
        orange = PatternFill("solid", start_color="FFA500")
        blue = PatternFill("solid", start_color="0000FF")
        green = PatternFill("solid", start_color="00FF00")
        gray = PatternFill("solid", start_color="D3D3D3")

        for i, col_name in enumerate(df_output.columns):
            if "Pct_OI_Percentile" not in col_name:
                continue
            letter = get_column_letter(i + 1)
            for r in range(2, len(df_output) + 2):
                cell = sheet[f"{letter}{r}"]
                try:
                    v = float(cell.value)
                    if v >= 80: cell.fill = red
                    elif 50 <= v < 80: cell.fill = orange
                    elif 20 <= v < 50: cell.fill = blue
                    elif v <= 20: cell.fill = green
                    else: cell.fill = gray
                except:
                    pass

        print(f"{sheet_name} sheet 已生成")
