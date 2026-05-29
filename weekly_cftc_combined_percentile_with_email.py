# merged_precious_metals_cftc_analysis_final_unified.py
# 最终优化版：净持仓占比分位 + 删除重复Historical列 + 删除Avg/Median OI + 完整当前分析

import akshare as ak
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import requests
import zipfile
import shutil
import warnings
import smtplib
import os
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from pathlib import Path
from datetime import datetime, timedelta
from matplotlib.dates import DateFormatter
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

# ====================== 配置 ======================
BASE_DIR = Path.cwd()
CFTC_XLS_FOLDER = BASE_DIR / "cftc_xls_all"
CFTC_ZIP_FOLDER = BASE_DIR / "cftc_zips"
OUTPUT_FOLDER = BASE_DIR / "cftc_plots_percentile"
OUTPUT_EXCEL = BASE_DIR / f"CFTC_Gold_Silver_Platinum_With_Prices_{datetime.now().strftime('%Y%m%d')}.xlsx"
BACKTEST_SHEET = "Backtest_Analysis"

for folder in [CFTC_XLS_FOLDER, CFTC_ZIP_FOLDER, OUTPUT_FOLDER]:
    folder.mkdir(parents=True, exist_ok=True)

BASE_URL = "https://www.cftc.gov/files/dea/history/"
HIST_ZIP = "com_disagg_xls_hist_2006_2016.zip"

METAL_FILTERS = {
    "Gold":     ["GOLD - COMMODITY EXCHANGE INC."],
    "Silver":   ["SILVER - COMMODITY EXCHANGE INC."],
    "Platinum": ["PLATINUM - NEW YORK MERCANTILE EXCHANGE"]
}

CATEGORIES = ["Producers", "Swap_Dealers", "Money_Managers", "Other_Reportables", "Non_Reportables"]

DATE_COL = 'Report_Date_as_MM_DD_YYYY'
OI_COL = 'Open_Interest_All'
MARKET_COL = 'Market_and_Exchange_Names'

TIME_PERIODS = {
    "1Y": timedelta(days=365),
    "3Y": timedelta(days=3*365),
    "5Y": timedelta(days=5*365),
    "10Y": timedelta(days=10*365)
}  # 已移除 Historical，避免重复

# 邮件配置（从环境变量读取，避免泄露）
EMAIL_FROM = os.environ.get('EMAIL_FROM', '')
EMAIL_TO = os.environ.get('EMAIL_TO', '')
EMAIL_PASSWORD = os.environ.get('EMAIL_PASSWORD', '')
SMTP_SERVER = os.environ.get('SMTP_SERVER', 'smtp.qq.com')
SMTP_PORT = int(os.environ.get('SMTP_PORT', '465'))

# ====================== 获取价格数据 ======================
def fetch_precious_metals_prices():
    print("AKShare 版本:", ak.__version__)
    price_dfs = {}
    column_rename = {'日期': 'date', '收盘价': 'close'}

    for symbol, metal in [("AU0", "Gold"), ("AG0", "Silver")]:
        try:
            df = ak.futures_main_sina(symbol=symbol)
            df = df.rename(columns=column_rename)
            if 'date' in df.columns and 'close' in df.columns:
                df['date'] = pd.to_datetime(df['date'])
                price_dfs[metal] = df[['date', 'close']].sort_values('date').set_index('date')
                print(f"✓ {metal} 价格数据加载成功")
        except Exception as e:
            print(f"{metal} price failed: {e}")

    try:
        df = ak.futures_foreign_hist(symbol="XPT")
        date_col = next((c for c in df.columns if 'date' in c.lower()), None)
        close_col = next((c for c in df.columns if 'close' in c.lower() or 'settle' in c.lower()), None)
        if date_col and close_col:
            df = df.rename(columns={date_col: 'date', close_col: 'close'})
            df['date'] = pd.to_datetime(df['date'])
            price_dfs['Platinum'] = df[['date', 'close']].sort_values('date').set_index('date')
            print(f"✓ Platinum 价格数据加载成功")
    except Exception as e:
        print(f"Platinum price failed: {e}")

    return price_dfs

# ====================== 下载 CFTC 数据 ======================
def download_and_unzip(url, zip_dir, extract_dir, year=None, force=False):
    filename = url.split("/")[-1]
    zip_path = zip_dir / filename
    if not force and (extract_dir / f"com_disagg_{year or 'hist'}.xls").exists():
        return
    if not zip_path.exists() or force:
        r = requests.get(url, stream=True, timeout=60)
        r.raise_for_status()
        with open(zip_path, "wb") as f:
            for chunk in r.iter_content(8192):
                f.write(chunk)
    with zipfile.ZipFile(zip_path) as z:
        for member in z.namelist():
            if member.lower().endswith((".xls", ".xlsx")):
                z.extract(member, extract_dir)
                if year:
                    new_path = extract_dir / f"com_disagg_{year}.xls"
                    shutil.move(extract_dir / member, new_path)

def download_all_cftc_data():
    download_and_unzip(BASE_URL + HIST_ZIP, CFTC_ZIP_FOLDER, CFTC_XLS_FOLDER)
    current_year = datetime.now().year
    for year in range(2017, current_year):
        download_and_unzip(BASE_URL + f"com_disagg_xls_{year}.zip", CFTC_ZIP_FOLDER, CFTC_XLS_FOLDER, year=year)
    download_and_unzip(BASE_URL + f"com_disagg_xls_{current_year}.zip", CFTC_ZIP_FOLDER, CFTC_XLS_FOLDER, year=current_year, force=True)

# ====================== 处理单个文件 ======================
def process_file(file):
    try:
        df = pd.read_excel(file, engine='xlrd')
        if DATE_COL not in df.columns or MARKET_COL not in df.columns:
            return None
        df['Report_Date'] = pd.to_datetime(df[DATE_COL], errors='coerce')
        df = df.dropna(subset=['Report_Date'])

        mappings = {
            "Producers": ("Prod_Merc_Positions_Long_ALL", "Prod_Merc_Positions_Short_ALL"),
            "Swap_Dealers": ("Swap_Positions_Long_All", "Swap__Positions_Short_All"),
            "Money_Managers": ("M_Money_Positions_Long_ALL", "M_Money_Positions_Short_ALL"),
            "Other_Reportables": ("Other_Rept_Positions_Long_ALL", "Other_Rept_Positions_Short_ALL"),
            "Non_Reportables": ("NonRept_Positions_Long_All", "NonRept_Positions_Short_All")
        }

        for cat, (l, s) in mappings.items():
            if l in df.columns and s in df.columns:
                df[f"{cat}_Long"] = df[l]
                df[f"{cat}_Short"] = df[s]
                df[f"{cat}_Net"] = df[l] - df[s]
                df[f"{cat}_Pct_OI"] = df[f"{cat}_Net"] / df[OI_COL] * 100

        return df
    except Exception as e:
        print(f"处理文件失败: {e}")
        return None

# ====================== 主 sheet 处理 ======================
def process_and_write(writer, price_dfs):
    all_data = {m: pd.DataFrame() for m in METAL_FILTERS}

    for file in sorted(CFTC_XLS_FOLDER.glob("*.xls")):
        df = process_file(file)
        if df is not None:
            for metal, keywords in METAL_FILTERS.items():
                mask = df[MARKET_COL].str.upper().isin([k.upper() for k in keywords])
                if mask.any():
                    all_data[metal] = pd.concat([all_data[metal], df[mask]]).copy()

    for metal, df in all_data.items():
        if df.empty:
            print(f"{metal} 数据为空，跳过处理")
            continue

        df = df.sort_values('Report_Date').drop_duplicates('Report_Date', keep='last').copy()
        df = df[df['Report_Date'] >= '2007-01-01']
        df['Report_Date'] = pd.to_datetime(df['Report_Date'])
        df = df.set_index('Report_Date').copy()

        if metal in price_dfs and not price_dfs[metal].empty:
            price_series = price_dfs[metal]['close'].sort_index()
            df['Tuesday_Date'] = df.index - pd.Timedelta(days=3)
            df['close'] = np.nan
            for idx, row in df.iterrows():
                tue_date = row['Tuesday_Date']
                if tue_date in price_series.index:
                    df.at[idx, 'close'] = price_series.loc[tue_date]
                else:
                    prev_dates = price_series.index[price_series.index < tue_date]
                    if not prev_dates.empty:
                        nearest_date = prev_dates.max()
                        df.at[idx, 'close'] = price_series.loc[nearest_date]
            df['close'] = df['close'].ffill().bfill()

            print_check = df.reset_index()
            print(f"\n{metal} 最近10周 close 匹配检查：")
            print(print_check[['Report_Date', 'Tuesday_Date', 'close']].tail(10))

        df = df.reset_index()

        # 全历史分位
        for cat in CATEGORIES:
            pct_col = f"{cat}_Pct_OI"
            if pct_col not in df.columns:
                continue
            lower = df[pct_col].quantile(0.01)
            upper = df[pct_col].quantile(0.99)
            clipped = df[pct_col].clip(lower, upper)
            df[f"{cat}_Pct_OI_Percentile"] = clipped.rank(pct=True) * 100

        # 滚动窗口分位（不含Historical）
        latest = df['Report_Date'].max()
        for p_name, delta in TIME_PERIODS.items():
            p_df = df[df['Report_Date'] >= latest - delta].copy()
            for cat in CATEGORIES:
                pct_col = f"{cat}_Pct_OI"
                if pct_col in p_df.columns:
                    lower = p_df[pct_col].quantile(0.01)
                    upper = p_df[pct_col].quantile(0.99)
                    clipped = p_df[pct_col].clip(lower, upper)
                    ranks = clipped.rank(pct=True) * 100
                    df[f"{cat}_{p_name}_Pct_OI_Percentile"] = np.nan
                    df.loc[p_df.index, f"{cat}_{p_name}_Pct_OI_Percentile"] = ranks

        # 输出列
        cols = ['Report_Date', OI_COL]
        for cat in CATEGORIES:
            cols += [f"{cat}_Long", f"{cat}_Short", f"{cat}_Net", f"{cat}_Pct_OI", f"{cat}_Pct_OI_Percentile"]
            for p in TIME_PERIODS:
                cols.append(f"{cat}_{p}_Pct_OI_Percentile")
        if 'close' in df.columns:
            cols.append('close')

        df_output = df[cols].sort_values('Report_Date', ascending=False)
        df_output.to_excel(writer, sheet_name=metal, index=False, float_format="%.2f")

        # 颜色填充
        sheet = writer.sheets[metal]
        red = PatternFill("solid", start_color="FF0000")
        orange = PatternFill("solid", start_color="FFA500")
        blue = PatternFill("solid", start_color="0000FF")
        green = PatternFill("solid", start_color="00FF00")
        gray = PatternFill("solid", start_color="D3D3D3")

        for i, col in enumerate(df_output.columns):
            if 'Pct_OI_Percentile' not in col:
                continue
            letter = get_column_letter(i + 1)
            for r in range(2, len(df_output) + 2):
                cell = sheet[f"{letter}{r}"]
                try:
                    v = float(cell.value)
                    if v >= 80: cell.fill = red
                    elif 50 <= v < 80: cell.fill = orange
                    elif 20 <= v < 50: cell.fill = blue
                    elif v <= 20: cell.fill = green
                    else: cell.fill = gray
                except:
                    pass

        print(f"{metal} 主 sheet 已生成")

    return all_data

# ====================== 2014-2020 sheet ======================
def add_2014_2020_sheets(writer, all_data, price_dfs):
    for metal, raw_df in all_data.items():
        if raw_df.empty:
            print(f"{metal} 2014-2020 数据为空，跳过")
            continue

        df = raw_df.sort_values('Report_Date').drop_duplicates('Report_Date', keep='last').copy()
        df = df[(df['Report_Date'] >= '2014-01-01') & (df['Report_Date'] <= '2020-12-31')]
        if len(df) < 30:
            print(f"{metal} 2014-2020 数据不足，跳过")
            continue

        df = df.set_index('Report_Date')

        if metal in price_dfs and not price_dfs[metal].empty:
            price_series = price_dfs[metal]['close'].sort_index()
            df['Tuesday_Date'] = df.index - pd.Timedelta(days=3)
            df['close'] = np.nan
            for idx, row in df.iterrows():
                tue_date = row['Tuesday_Date']
                if tue_date in price_series.index:
                    df.at[idx, 'close'] = price_series.loc[tue_date]
                else:
                    prev_dates = price_series.index[price_series.index < tue_date]
                    if not prev_dates.empty:
                        nearest_date = prev_dates.max()
                        df.at[idx, 'close'] = price_series.loc[nearest_date]
            df['close'] = df['close'].ffill().bfill()

        df = df.reset_index()

        for cat in CATEGORIES:
            pct_col = f"{cat}_Pct_OI"
            if pct_col not in df.columns:
                continue
            lower = df[pct_col].quantile(0.01)
            upper = df[pct_col].quantile(0.99)
            clipped = df[pct_col].clip(lower, upper)
            df[f"{cat}_Pct_OI_Percentile"] = clipped.rank(pct=True) * 100

        if 'close' in df.columns and df['close'].notna().sum() > 1:
            df['Weekly_Return'] = df['close'].pct_change(1) * 100
            df['Next_Weekly_Return'] = df['close'].pct_change(1).shift(-1) * 100
        else:
            df['Weekly_Return'] = np.nan
            df['Next_Weekly_Return'] = np.nan

        for cat in CATEGORIES:
            pct_col = f"{cat}_Pct_OI_Percentile"
            if pct_col not in df.columns:
                continue
            df[f'{cat}_Signal'] = '中性'
            df.loc[df[pct_col] >= 80, f'{cat}_Signal'] = '极度拥挤多头'
            df.loc[(df[pct_col] >= 50) & (df[pct_col] < 80), f'{cat}_Signal'] = '强势多头 / 拥挤多头'
            df.loc[(df[pct_col] >= 20) & (df[pct_col] < 50), f'{cat}_Signal'] = '弱势空头 / 拥挤空头'
            df.loc[df[pct_col] < 20, f'{cat}_Signal'] = '极度拥挤空头'

            df[f'{cat}_Week_Match'] = 0
            df.loc[(df[pct_col] >= 50) & (df['Weekly_Return'] > 0), f'{cat}_Week_Match'] = 1
            df.loc[(df[pct_col] < 50) & (df['Weekly_Return'] < 0), f'{cat}_Week_Match'] = 1

            df[f'{cat}_Next_Match'] = 0
            df.loc[(df[pct_col] >= 50) & (df['Next_Weekly_Return'] > 0), f'{cat}_Next_Match'] = 1
            df.loc[(df[pct_col] < 50) & (df['Next_Weekly_Return'] < 0), f'{cat}_Next_Match'] = 1

        cols = ['Report_Date', OI_COL]
        if 'close' in df.columns:
            cols.append('close')
        for cat in CATEGORIES:
            cols += [f"{cat}_Long", f"{cat}_Short", f"{cat}_Net", f"{cat}_Pct_OI",
                     f"{cat}_Pct_OI_Percentile", f"{cat}_Signal", 
                     f"{cat}_Week_Match", f"{cat}_Next_Match"]
        cols += ['Weekly_Return', 'Next_Weekly_Return']

        df_output = df[cols].sort_values('Report_Date', ascending=False)
        sheet_name = f"{metal}_2014_2020"
        df_output.to_excel(writer, sheet_name=sheet_name, index=False, float_format="%.2f")

        # 颜色标记
        sheet = writer.sheets[sheet_name]
        red = PatternFill("solid", start_color="FF0000")
        orange = PatternFill("solid", start_color="FFA500")
        blue = PatternFill("solid", start_color="0000FF")
        green = PatternFill("solid", start_color="00FF00")
        gray = PatternFill("solid", start_color="D3D3D3")

        for i, col_name in enumerate(df_output.columns):
            if "Pct_OI_Percentile" not in col_name:
                continue
            letter = get_column_letter(i + 1)
            for r in range(2, len(df_output) + 2):
                cell = sheet[f"{letter}{r}"]
                try:
                    v = float(cell.value)
                    if v >= 80: cell.fill = red
                    elif 50 <= v < 80: cell.fill = orange
                    elif 20 <= v < 50: cell.fill = blue
                    elif v <= 20: cell.fill = green
                    else: cell.fill = gray
                except:
                    pass

        print(f"{sheet_name} sheet 已生成")

# ====================== 图表生成、回溯、当前分析、邮件（保持最新逻辑） ======================
def draw_percentile_chart(metal, df, title_suffix, filename_suffix):
    fig, axes = plt.subplots(6, 1, figsize=(15, 26), sharex=True, gridspec_kw={'height_ratios': [1]*6})
    fig.suptitle(f"{metal} Percentiles & Spot Price {title_suffix}", fontsize=16, fontweight='bold')
    ax_price = axes[0]
    if 'close' in df.columns and not df['close'].isna().all():
        ax_price.plot(df['Report_Date'], df['close'], color='black', lw=1.5)
        ax_price.set_ylabel('Spot Price')
        ax_price.grid(True, alpha=0.3)
    for i, cat in enumerate(CATEGORIES):
        ax = axes[i + 1]
        cols_to_try = [f"{cat}_Pct_OI_Percentile", f"{cat}_Norm_Avg_Percentile", f"{cat}_Historical_Percentile"]
        col = next((c for c in cols_to_try if c in df.columns and not df[c].isna().all()), None)
        if col is None:
            ax.text(0.5, 50, "No data", ha='center', va='center')
            continue
        colors = ['darkred' if v >= 80 else 'orange' if 50 <= v < 80 else 'blue' if 20 <= v < 50 else 'darkgreen' for v in df[col]]
        ax.scatter(df['Report_Date'], df[col], c=colors, s=40, alpha=0.8)
        ax.axhline(80, color="darkred", ls=":")
        ax.axhline(50, color="orange", ls=":")
        ax.axhline(20, color="blue", ls=":")
        ax.axhline(50, color="gray", ls="--")
        ax.set_ylabel(f"{cat} (%)")
        ax.grid(True, alpha=0.3)
        ax.set_ylim(0, 100)
    legend_elements = [mpatches.Patch(color='darkred', label='≥80%'), mpatches.Patch(color='orange', label='50–79%'),
                       mpatches.Patch(color='blue', label='20–49%'), mpatches.Patch(color='darkgreen', label='≤20%')]
    fig.legend(handles=legend_elements, loc='lower center', bbox_to_anchor=(0.5, -0.02), ncol=1, fontsize=9)
    axes[-1].xaxis.set_major_formatter(DateFormatter('%Y-%m'))
    plt.xticks(rotation=45)
    fig.tight_layout(rect=[0, 0.05, 1, 0.92])
    plt.savefig(OUTPUT_FOLDER / f"{metal}_{filename_suffix}.png", dpi=180, bbox_inches='tight')
    plt.close(fig)

def generate_plots():
    if not OUTPUT_EXCEL.exists():
        print("Excel 文件不存在，跳过图表生成")
        return
    xl = pd.ExcelFile(OUTPUT_EXCEL, engine='openpyxl')
    for metal in ["Gold", "Silver", "Platinum"]:
        if metal not in xl.sheet_names:
            continue
        df = xl.parse(metal)
        df['Report_Date'] = pd.to_datetime(df['Report_Date'], errors='coerce')
        df = df.dropna(subset=['Report_Date']).sort_values('Report_Date')
        if df.empty:
            continue
        latest_date = df['Report_Date'].max()
        draw_percentile_chart(metal, df, "(Full History)", "Historical_Percentiles")
        sheet_name = f"{metal}_2014_2020"
        if sheet_name in xl.sheet_names:
            df_1420 = xl.parse(sheet_name)
            df_1420['Report_Date'] = pd.to_datetime(df_1420['Report_Date'], errors='coerce')
            df_1420 = df_1420.dropna(subset=['Report_Date']).sort_values('Report_Date')
            df_1420 = df_1420[(df_1420['Report_Date'] >= '2014-01-01') & (df_1420['Report_Date'] <= '2020-12-31')]
            if not df_1420.empty:
                draw_percentile_chart(metal, df_1420, "(2014-2020)", "2014_2020_Historical_Percentiles")
        for period in ["1Y", "3Y", "5Y", "10Y"]:
            period_start = latest_date - TIME_PERIODS[period]
            period_df = df[df['Report_Date'] >= period_start].copy()
            if len(period_df) < 10:
                continue
            pcols = [f"{cat}_{period}_Pct_OI_Percentile" for cat in CATEGORIES]
            if all(c not in period_df.columns or period_df[c].isna().all() for c in pcols):
                continue
            draw_percentile_chart(metal, period_df, f"({period})", f"{period}_Percentiles")
    print("图表生成完成")

def backtest_analysis():
    xl = pd.ExcelFile(OUTPUT_EXCEL, engine='openpyxl')
    results = []
    for metal in METAL_FILTERS:
        if metal not in xl.sheet_names: continue
        df = xl.parse(metal)
        df['Report_Date'] = pd.to_datetime(df['Report_Date'])
        df = df.sort_values('Report_Date').set_index('Report_Date')
        pct_col = 'Money_Managers_Pct_OI_Percentile'
        if pct_col not in df.columns or 'close' not in df.columns: continue
        df['Overbought'] = (df[pct_col] >= 80).astype(int)
        df['Oversold'] = (df[pct_col] <= 20).astype(int)
        df['Overbought_Signal'] = df['Overbought'].rolling(2).sum() >= 2
        df['Oversold_Signal'] = df['Oversold'].rolling(2).sum() >= 2
        for h in [1, 2, 4]:
            df[f'Return_{h}w'] = df['close'].pct_change(h).shift(-h) * 100
        for sig, name in [('Overbought_Signal', '做多拥挤 (>=80%)'), ('Oversold_Signal', '做空拥挤 (<=20%)')]:
            instances = df[df[sig]]
            if instances.empty: continue
            stats = {'Metal': metal, 'Signal_Type': name, 'Count': len(instances)}
            for h in [1, 2, 4]:
                rets = instances[f'Return_{h}w'].dropna()
                stats[f'Avg_Return_{h}w'] = rets.mean() if not rets.empty else np.nan
                stats[f'Win_Rate_{h}w'] = (rets > 0).mean() * 100 if not rets.empty else np.nan
            results.append(stats)
    if results:
        result_df = pd.DataFrame(results)
        with pd.ExcelWriter(OUTPUT_EXCEL, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            result_df.to_excel(writer, sheet_name=BACKTEST_SHEET, index=False)
        print("回溯分析已写入")

# ====================== 当前历史分位分析（优化版） ======================
def current_signal_analysis():
    xl = pd.ExcelFile(OUTPUT_EXCEL, engine='openpyxl')
    print("\n=== 当前历史分位分析（最新报告周） ===\n")
    analysis_lines = []

    for metal in ["Gold", "Silver", "Platinum"]:
        if metal not in xl.sheet_names:
            continue

        full_df = xl.parse(metal)

        print(f"→ 正在读取 {metal} sheet，数据量: {len(full_df)} 行")   # 调试信息

        full_df['Report_Date'] = pd.to_datetime(full_df['Report_Date'], errors='coerce')
        full_df = full_df.dropna(subset=['Report_Date', 'Money_Managers_Pct_OI_Percentile'])

        latest_row = full_df.sort_values('Report_Date', ascending=False).head(1)
        pct = latest_row['Money_Managers_Pct_OI_Percentile'].iloc[0]
        latest_date = latest_row['Report_Date'].iloc[0]

        # 各区间统计
        count_ge80 = len(full_df[full_df['Money_Managers_Pct_OI_Percentile'] >= 80])
        count_le20 = len(full_df[full_df['Money_Managers_Pct_OI_Percentile'] <= 20])
        count_50_80 = len(full_df[(full_df['Money_Managers_Pct_OI_Percentile'] >= 50) & 
                                  (full_df['Money_Managers_Pct_OI_Percentile'] < 80)])
        count_20_50 = len(full_df[(full_df['Money_Managers_Pct_OI_Percentile'] > 20) & 
                                  (full_df['Money_Managers_Pct_OI_Percentile'] < 50)])
        # 出现次数统计
        total_weeks = len(full_df)
        print(f"\n{metal} 总历史周数: {total_weeks} 周")
        print(f"  >=80% : {count_ge80} 次")
        print(f"  50-80%: {count_50_80} 次")
        print(f"  20-50%: {count_20_50} 次")
        print(f"  <=20% : {count_le20} 次")

        if pct >= 80:
            msg = (f"{metal} 当前处于**极度拥挤多头** ({pct:.1f}%) - 日期: {latest_date.date()}\n"
                   f"历史>=80%出现次数: {count_ge80} 次\n"
                   f"专业建议：市场情绪极端乐观，历史上此类信号后回调或反转概率较高。\n"
                   f"**推荐操作**：偏谨慎看空，建议减持多头或建立空头头寸，设置严格止损。")

        elif pct <= 20:
            msg = (f"{metal} 当前处于**极度拥挤空头** ({pct:.1f}%) - 日期: {latest_date.date()}\n"
                   f"历史<=20%出现次数: {count_le20} 次\n"
                   f"专业建议：市场情绪极端悲观，历史上此类信号后反弹或反转概率较高。\n"
                   f"**推荐操作**：偏向看多，可逐步建立或加仓多头头寸，注意止损防范进一步下探。")

        elif 50 <= pct < 80:
            msg = (f"{metal} 当前处于**强势多头区间** ({pct:.1f}%) - 日期: {latest_date.date()}\n"
                   f"历史50-80%出现次数: {count_50_80} 次 | 20-50%出现次数: {count_20_50} 次\n"
                   f"专业建议：市场情绪偏强，趋势可能延续，但需关注分位高度。")

            if pct >= 70:
                msg += "\n**推荐操作**：趋势跟随看多，但分位已较高，注意过热风险，设置 trailing stop 保护利润，避免追高。"
            else:
                msg += "\n**推荐操作**：可趋势跟随看多，风险相对可控，回调时可考虑加仓。"

        elif 20 < pct < 50:
            msg = (f"{metal} 当前处于**弱势空头区间** ({pct:.1f}%) - 日期: {latest_date.date()}\n"
                   f"历史20-50%出现次数: {count_20_50} 次\n"
                   f"专业建议：市场情绪偏弱，短期仍有下行压力，但反弹机会也在增加。\n"
                   f"**推荐操作**：保持中性或轻仓做空，等待更极端信号或技术支撑位再加大仓位。")

        else:
            msg = (f"{metal} 当前处于**中性区间** ({pct:.1f}%) - 日期: {latest_date.date()}\n"
                   f"历史中性区间出现次数: {len(full_df) - count_ge80 - count_le20 - count_50_80 - count_20_50} 次\n"
                   f"专业建议：市场方向性不强，建议观望为主。")

        print(msg + "\n")
        analysis_lines.append(msg)

    return "\n\n".join(analysis_lines)

def send_email(current_analysis=""):
    msg = MIMEMultipart()
    msg['From'] = EMAIL_FROM
    msg['To'] = EMAIL_TO
    msg['Subject'] = f'CFTC 数据更新 {datetime.now().strftime("%Y-%m-%d")}'
    body = f"""CFTC 数据更新完成。\n\n=== 当前分析 ===\n{current_analysis}\n祝交易顺利！"""
    msg.attach(MIMEText(body, 'plain', 'utf-8'))

    if OUTPUT_EXCEL.exists():
        with open(OUTPUT_EXCEL, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f"attachment; filename={OUTPUT_EXCEL.name}")
            msg.attach(part)
    for png in OUTPUT_FOLDER.glob("*.png"):
        with open(png, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f"attachment; filename={png.name}")
            msg.attach(part)

    try:
        server = smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT)
        server.login(EMAIL_FROM, EMAIL_PASSWORD)
        server.send_message(msg)
        server.quit()
        print("邮件已发送")
    except Exception as e:
        print(f"邮件发送失败: {e}")

# ====================== 主程序 ======================
if __name__ == "__main__":
    print(f"\n=== 开始 {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ===\n")
    price_dfs = fetch_precious_metals_prices()
    current_year = datetime.now().year
    (CFTC_ZIP_FOLDER / f"com_disagg_xls_{current_year}.zip").unlink(missing_ok=True)
    download_all_cftc_data()

    with pd.ExcelWriter(OUTPUT_EXCEL, engine='openpyxl') as writer:
        all_data = process_and_write(writer, price_dfs)
        add_2014_2020_sheets(writer, all_data, price_dfs)

    generate_plots()
    backtest_analysis()
    current_analysis = current_signal_analysis()
    send_email(current_analysis)

    print("\n=== 完成 ===\n")
